from django.http import HttpResponse
from django.http import JsonResponse
from django.template import loader
from .models import Member
from .utility import coutMember, getColor, fetch_data, updateDB

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


def dowload_exel(request):
    wb = Workbook()
    ws = wb.active

    headers = [
        "รหัสโครงการ",
        "ชื่อโครงการ",
        "หน่วยงาน",
        "สถานะ",
        "ตัวแบ่งสถานะ",
        "ยอดโครงการ(บาท)",
        "เบิกเงิน(บาท)",
        "หมายเหตุ",
    ]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header
        ws[f"{col_letter}1"].font = Font(bold=True)

    members = Member.objects.all()

    for row_num, member in enumerate(members, 2):
        ws[f"A{row_num}"] = member.code
        ws[f"B{row_num}"] = member.projectName
        ws[f"C{row_num}"] = member.agency
        ws[f"D{row_num}"] = member.status
        ws[f"E{row_num}"] = member.detail
        ws[f"F{row_num}"] = f"{member.total:,.2f}"
        ws[f"G{row_num}"] = f"{member.withdraw:,.2f}"
        ws[f"H{row_num}"] = member.note

    font = Font(name="TH Sarabun New", size=14)
    for col in ws.iter_cols(min_col=1, max_col=8, min_row=1):
        for cell in col:
            cell.font = font

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet; charset=utf-8"
    )
    response["Content-Disposition"] = "attachment; filename=project_data.xlsx"
    wb.save(response)
    print("dowload!!!")
    return response


def search_view(request):
    template = loader.get_template("home.html")

    query = request.GET.get("search", "").strip()
    result = Member.search(query=query)
    result = getColor(result)

    mymembers = Member.objects.all().values()
    percent = coutMember(mymembers)

    context = {"result": result, "query": query, "percents": percent}
    return HttpResponse(template.render(context, request))


def finance_data(request, member_id):
    try:
        member = Member.objects.get(id=member_id)
        remain = member.total - member.withdraw

        data = {
            "total": f"{member.total:,.2f}",
            "withdraw": member.withdraw,
            "remain": remain,
        }

        response = JsonResponse(data)
        response["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"

        return response

    except Member.DoesNotExist:
        return HttpResponse(status=404)


def home(request):
    (
        code_list,
        name_list,
        agency_list,
        status_list,
        total_list,
        withdraw_list,
        detail_list,
    ) = fetch_data()

    updateDB(
        code_list,
        name_list,
        agency_list,
        status_list,
        total_list,
        withdraw_list,
        detail_list,
    )

    mymembers = Member.objects.all().values()
    mymembers = getColor(mymembers)
    percent = coutMember(mymembers)

    template = loader.get_template("home.html")
    context = {"mymembers": mymembers, "percents": percent}
    return HttpResponse(template.render(context, request))
